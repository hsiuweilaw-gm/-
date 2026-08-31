# -*- coding: utf-8 -*-
"""端對端測試：實際產生 docx / xlsx / pdf 檔並執行 CLI 遮罩。"""
import sys
import zipfile
from pathlib import Path

import pytest

SAMPLES_DIR = Path(__file__).parent.parent / "samples"
sys.path.insert(0, str(SAMPLES_DIR))

import make_samples  # noqa: E402
from tw_pii_masker.cli import run  # noqa: E402

FAKE = make_samples.FAKE


@pytest.fixture()
def sample_dir(tmp_path):
    make_samples.make_docx(tmp_path)
    make_samples.make_xlsx(tmp_path)
    make_samples.make_pdf(tmp_path)
    return tmp_path


def _docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as z:
        return z.read("word/document.xml").decode("utf-8")


def test_docx_masking(sample_dir):
    src = sample_dir / "示範_理賠申請書.docx"
    rc = run([str(src)])
    assert rc == 0
    out = sample_dir / "示範_理賠申請書_masked.docx"
    assert out.exists()
    xml = _docx_text(out)
    for key in ("id", "mobile", "email", "card"):
        assert FAKE[key] not in xml, key
    assert "王小明" not in xml
    assert "和平東路" not in xml
    # 部分遮罩應保留可比對的頭尾
    assert "A12" in xml
    assert "王○○" in xml
    # 原始檔不得被修改
    assert FAKE["id"] in _docx_text(src)


def test_xlsx_masking(sample_dir):
    from openpyxl import load_workbook
    src = sample_dir / "示範_客戶名單.xlsx"
    rc = run([str(src)])
    assert rc == 0
    out = sample_dir / "示範_客戶名單_masked.xlsx"
    wb = load_workbook(out)
    ws = wb["客戶名單"]
    values = " | ".join(str(c.value) for row in ws.iter_rows() for c in row)
    assert FAKE["id"] not in values
    assert "0912-345-678" not in values
    assert "王小明" not in values
    assert "04595257" not in values          # 統編（有「統一編號」前後文）
    assert "李大華" not in values            # 要保人欄位觸發姓名偵測
    assert "A12" in values                   # 部分遮罩保留頭
    # 表頭「姓名」「身分證字號」等欄位名稱不應被遮罩
    assert "姓名" in values


def test_pdf_masking(sample_dir):
    import pymupdf
    src = sample_dir / "示範_契約通知書.pdf"
    rc = run([str(src)])
    assert rc == 0
    out = sample_dir / "示範_契約通知書_masked.pdf"
    doc = pymupdf.open(out)
    text = "".join(page.get_text() for page in doc)
    doc.close()
    for key in ("id", "mobile", "email"):
        assert FAKE[key] not in text, key
    assert "王小明" not in text
    # redaction 為真實移除：原文以文字搜尋也找不到
    assert "A123456789" not in text


def test_dry_run_outputs_nothing(sample_dir):
    src = sample_dir / "示範_理賠申請書.docx"
    rc = run([str(src), "--dry-run"])
    assert rc == 0
    assert not (sample_dir / "示範_理賠申請書_masked.docx").exists()


def test_report_json(sample_dir, tmp_path):
    import json
    src = sample_dir / "示範_理賠申請書.docx"
    report_file = tmp_path / "report.json"
    rc = run([str(src), "--report", str(report_file)])
    assert rc == 0
    data = json.loads(report_file.read_text(encoding="utf-8"))
    findings = data["files"][0]["findings"]
    assert findings
    # 預設不得寫入原始個資
    assert all("original" not in f for f in findings)
    assert data["files"][0]["counts"].get("身分證字號", 0) >= 1


def test_label_mode(sample_dir):
    src = sample_dir / "示範_理賠申請書.docx"
    rc = run([str(src), "-m", "label", "--suffix", "_label"])
    assert rc == 0
    xml = _docx_text(sample_dir / "示範_理賠申請書_label.docx")
    assert "[身分證字號]" in xml


def test_txt_masking(tmp_path):
    src = tmp_path / "note.txt"
    src.write_text("客戶王先生 身分證 A123456789 電話0912345678", encoding="utf-8")
    rc = run([str(src)])
    assert rc == 0
    out = (tmp_path / "note_masked.txt").read_text(encoding="utf-8")
    assert "A123456789" not in out
    assert "0912345678" not in out


def test_xlsx_numeric_phone_and_bare_address(tmp_path):
    from openpyxl import Workbook, load_workbook
    src = tmp_path / "名單.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "行動電話", "地址"])
    ws.append(["王小明", 912345678, "板橋區文化路一段23號5樓"])  # 電話存成數值、地址未含縣市
    wb.save(src)
    rc = run([str(src)])
    assert rc == 0
    out = load_workbook(tmp_path / "名單_masked.xlsx").active
    values = " | ".join(str(c.value) for row in out.iter_rows() for c in row)
    assert "912345678" not in values
    assert "0912****78" in values      # 補回開頭的 0 再遮罩
    assert "文化路" not in values
    assert "板橋區" in values


def test_xlsx_datetime_birthday_and_new_columns(tmp_path):
    import datetime
    from openpyxl import Workbook, load_workbook
    src = tmp_path / "名冊.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "生日 (YYYY-MM-DD)", "推介者名稱", "壽險證照號碼", "戶籍地址"])
    ws.append(["陶柏勲", datetime.date(1975, 3, 12), "葛昇威", "B103900033", "745號"])
    wb.save(src)
    rc = run([str(src)])
    assert rc == 0
    out = load_workbook(tmp_path / "名冊_masked.xlsx").active
    values = " | ".join(str(c.value) for row in out.iter_rows() for c in row)
    assert "陶柏勲" not in values and "陶○○" in values
    assert "1975" not in values and "****-**-**" in values
    assert "葛昇威" not in values
    assert "B103900033" not in values
    assert "745號" not in values
    # 表頭不得被遮
    assert "姓名" in values and "壽險證照號碼" in values


def test_xlsx_consistency_masking_in_free_text(tmp_path):
    """自由文字欄：同一姓名在無標籤處也要遮（原本只遮有標籤的第一次）。"""
    from openpyxl import Workbook, load_workbook
    src = tmp_path / "查核.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["系統編號", "標題", "說明"])
    ws.append([
        6819,
        "受理 遠雄人壽 保單 1150727LN00008 保戶 饒書寧 的 聯絡資料 相同",
        "要保人 饒書寧(F228969519):地址 新北市汐止區秀山路147巷22號 "
        "與 帳號 饒培杰(A121775567) 之 地址 相同",
    ])
    wb.save(src)
    rc = run([str(src)])
    assert rc == 0
    out = load_workbook(tmp_path / "查核_masked.xlsx").active
    values = " | ".join(str(c.value) for row in out.iter_rows() for c in row)
    assert "饒書寧" not in values       # 保戶姓名：兩欄多處全遮
    assert "饒培杰" in values           # 「帳號」後的業務員姓名保留
    assert "F228969519" not in values
    assert "秀山路" not in values
    assert "新北市汐止區" in values     # 行政區層級保留
    assert "系統編號" in values         # 表頭不動


def test_xlsx_learned_names_do_not_leak_across_files(tmp_path):
    """一致性遮罩以單一文件為範圍，不可跨檔案累積。"""
    from openpyxl import Workbook, load_workbook
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    for path, rows in ((a, ["要保人 周洺禾 申請"]), (b, ["營業處所 周洺禾路 施工"])):
        wb = Workbook()
        wb.active.append(["說明"])
        wb.active.append(rows)
        wb.save(path)
    rc = run([str(a), str(b)])
    assert rc == 0
    out_b = load_workbook(tmp_path / "b_masked.xlsx").active
    values = " | ".join(str(c.value) for row in out_b.iter_rows() for c in row)
    assert "周洺禾路" in values          # a 檔學到的姓名不得帶到 b 檔


def test_cli_mask_policy_no_flag(tmp_path):
    from openpyxl import Workbook, load_workbook
    src = tmp_path / "案件.xlsx"
    wb = Workbook()
    wb.active.append(["說明"])
    wb.active.append(["受理 遠雄人壽 保單 1150727LN00008 保戶 王小明"])
    wb.save(src)

    assert run([str(src)]) == 0
    values = " | ".join(str(c.value) for row in
                        load_workbook(tmp_path / "案件_masked.xlsx").active.iter_rows()
                        for c in row)
    assert "1150727LN00008" in values     # 預設保留保單號碼
    assert "王小明" not in values          # 姓名仍要遮

    assert run([str(src), "--mask-policy-no", "--suffix", "_p"]) == 0
    values = " | ".join(str(c.value) for row in
                        load_workbook(tmp_path / "案件_p.xlsx").active.iter_rows()
                        for c in row)
    assert "1150727LN00008" not in values  # 明確開啟後才遮


def test_xlsx_agent_names_kept_customer_names_masked(tmp_path):
    """業務員／帳號姓名保留、客戶姓名遮罩；鄰欄提到業務員不得放行客戶姓名。"""
    from openpyxl import Workbook, load_workbook
    src = tmp_path / "查核案件.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["系統編號", "標題", "說明"])
    ws.append([
        6820,
        "(業務侵佔) 受理 全球人壽 保戶 周洺禾 的 聯絡資料 與 業務員/營業處所 相同",
        "要保人 周洺禾(P123732535):地址 彰化縣線西鄉中華路121號 "
        "與 帳號 劉湘妘(N224463647) 之 地址 相同",
    ])
    wb.save(src)
    assert run([str(src)]) == 0
    values = " | ".join(str(c.value) for row in
                        load_workbook(tmp_path / "查核案件_masked.xlsx").active.iter_rows()
                        for c in row)
    assert "周洺禾" not in values and "周○○" in values   # 保戶：兩欄都遮
    assert "劉湘妘" in values                             # 業務員：保留
    assert "P123732535" not in values                    # 身分證仍遮
    assert "N224463647" not in values                    # 業務員身分證也遮
    assert "中華路" not in values                        # 地址仍遮

    # 加旗標可回復為一律遮罩
    assert run([str(src), "--mask-agent-names", "--suffix", "_a"]) == 0
    values = " | ".join(str(c.value) for row in
                        load_workbook(tmp_path / "查核案件_a.xlsx").active.iter_rows()
                        for c in row)
    assert "劉湘妘" not in values


def test_xlsx_label_decides_masking_per_occurrence(tmp_path):
    """客戶身分全文件優先；純內部人員（經手人）則各處保留。"""
    from openpyxl import Workbook, load_workbook
    src = tmp_path / "查核.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["系統編號", "標題", "說明"])
    ws.append([
        6826,
        "(業務侵佔) 受理 安聯人壽 保戶 林宥慈 的 聯絡資料 與 業務員/營業處所 相同",
        "要保人 林宥慈(F220755862):手機 0920094377 "
        "與 帳號 林宥慈(F220755862) 之 手機 相同",
    ])
    ws.append([6827, "經手人 葉珍玲(HC-20) 不具資格",
               "葉珍玲(HC-20)(hcv2223195) 的 壽險 與 公平"])
    wb.save(src)
    assert run([str(src)]) == 0
    out = load_workbook(tmp_path / "查核_masked.xlsx").active
    title, note = out["B2"].value, out["C2"].value
    assert "林○○" in title and "林宥慈" not in title    # 保戶 → 遮
    # 客戶身分優先：帳號處也一併遮，否則同一組身分證尾碼會把遮罩解開
    assert note.count("林○○") == 2
    assert "林宥慈" not in note
    # 經手人：兩欄都保留（含無標籤的那一處）
    assert "葉珍玲" in out["B3"].value and "葉珍玲" in out["C3"].value
    assert "hcv2223195" not in out["C3"].value   # 但帳號字串仍遮


def test_xlsx_customer_identity_not_leaked_via_agent_slot(tmp_path):
    """保戶漏遮的兩種實際情形（v1.6.0 修正）：

    1. 姓氏不在字庫（陽曼蘭）→ 客戶標籤本身即足以辨識
    2. 同一人在「帳號」處保留姓名，配上相同身分證尾碼，
       等於把「要保人」處的遮罩解開
    """
    from openpyxl import Workbook, load_workbook
    src = tmp_path / "查核.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["系統編號", "標題", "說明"])
    ws.append([7,
               "(業務侵佔) 受理 宏泰人壽 保戶 饒書寧 的 聯絡資料 與 業務員/營業處所 相同",
               "要保人 饒書寧(F228969519):地址 新北市汐止區秀山路147巷22號 "
               "與 帳號 饒培杰(A121775567) 之 地址 相同 "
               "地址 新北市汐止區秀山路147巷22號 與 帳號 饒書寧(F228969519) 之 地址 相同"])
    ws.append([10,
               "(業務侵佔) 受理 華南產物 保戶 陽曼蘭 的 聯絡資料 與 業務員/營業處所 相同",
               "主被保人 陽曼蘭(H291234795):手機 0913456769 "
               "與 帳號 林岳(H123456535) 之 電話 相同"])
    ws.append([18, "(銷售資格) 受理 明台產物 經手人 黃熙予 不具險種銷售資格",
               "明台產物_住宅火險(A-F01)<br>黃熙予(hcv1234535) 的 高齡"])
    ws.append([22, None, "業務員個人保件，饒書寧與饒培杰為直系親屬關係"])
    wb.save(src)
    assert run([str(src)]) == 0
    out = load_workbook(tmp_path / "查核_masked.xlsx").active
    values = " | ".join(str(c.value) for row in out.iter_rows() for c in row)

    # 客戶姓名：所有出現處都要遮，含「帳號」處與無標籤的備註
    assert "饒書寧" not in values
    assert "陽曼蘭" not in values and "陽○○" in values   # 姓氏不在字庫也要遮
    # 純內部人員：各處保留
    assert "饒培杰" in values
    assert "林岳" in values
    assert "黃熙予" in values                             # 經手人，含無標籤處
    # 其他個資照遮
    assert "F228969519" not in values and "秀山路" not in values
    assert "0913456769" not in values
