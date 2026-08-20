# -*- coding: utf-8 -*-
"""Excel (.xlsx / .xlsm) 處理器。

逐工作表、逐儲存格掃描文字內容並取代；.xlsm 會保留巨集。
同時預設清除活頁簿屬性中的作者資訊。

已知限制：
  - 舊版 .xls 不支援（請先以 Excel 另存為 .xlsx）
  - 只處理「文字」儲存格；以純數字形態儲存的號碼（如電話被存成數值 912345678）
    不會被辨識，建議先將該欄設為文字格式
  - 公式本身不會被改寫（避免破壞計算），但公式「顯示的結果」若含個資，
    會在報告中提出警告
  - 圖表、圖片、註解內的文字無法處理
"""
from __future__ import annotations

import datetime
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..engine import MaskedItem, MaskingEngine
from ..masking import mask_digits_keep
from ..report import Report

# 儲存格為「數值」但欄位提示顯示是電話時，補回開頭的 0 並遮罩
_PHONE_HINT_RE = re.compile(r"電話|手機|行動|傳真|TEL|Tel|tel|Phone|phone|Mobile|mobile|FAX|Fax|fax")


def mask_xlsx(input_path: str, output_path: str,
              engine: MaskingEngine, report: Report,
              keep_metadata: bool = False) -> None:
    engine.reset_learned()   # 一致性遮罩以單一文件為範圍
    keep_vba = str(input_path).lower().endswith(".xlsm")
    wb = load_workbook(input_path, keep_vba=keep_vba)

    # 第一遍：學習整份活頁簿中的姓名，供全文件一致遮罩
    # （自由文字中同一姓名往往只有部分出現處帶標籤）
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v or v.startswith("="):
                    continue
                hints = []
                if cell.row > 1:
                    hints.append(ws.cell(row=1, column=cell.column).value)
                if cell.column > 1:
                    hints.append(ws.cell(row=cell.row, column=cell.column - 1).value)
                engine.learn(v, "\n".join(
                    h for h in hints if isinstance(h, str) and h))

    for ws in wb.worksheets:
        # 工作表名稱也可能含姓名等個資
        new_title, items = engine.mask_text(ws.title)
        if items:
            report.add_all("工作表名稱「%s」" % new_title, items)
            ws.title = new_title

        formula_hits = 0
        # 提示須取自「原始值」快照：若讀取已遮罩的儲存格（例如表頭或左欄
        # 先被處理），提示會失真，導致後續儲存格誤判
        original = {}
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value:
                    original[(cell.row, cell.column)] = cell.value

        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                # 前後文提示：同欄表頭（第 1 列）、左側與上方儲存格，
                # 讓「姓名」欄底下的純姓名、表單式左標籤右值都能被判斷
                hints = []
                if cell.row > 1:
                    hints.append(original.get((1, cell.column)))
                    hints.append(original.get((cell.row - 1, cell.column)))
                if cell.column > 1:
                    hints.append(original.get((cell.row, cell.column - 1)))
                hint = "\n".join(h for h in hints if h)
                loc = "工作表「%s」儲存格 %s%d" % (
                    ws.title, get_column_letter(cell.column), cell.row)

                # 生日欄若是「日期型別」（非文字），原本不會被掃到，
                # 依欄位提示遮罩
                if isinstance(v, (datetime.datetime, datetime.date)):
                    if re.search(r"生日|出生", hint):
                        cell.value = "****-**-**"
                        report.add(loc, MaskedItem(
                            "birthdate", "出生日期", str(v)[:10], "****-**-**"))
                    continue

                # 電話被存成「數值」時開頭的 0 會消失（如 912345678），
                # 依欄位提示補回 0 後遮罩，並轉為文字避免再度流失
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    if float(v).is_integer() and v > 0:
                        digits = str(int(v))
                        if 8 <= len(digits) <= 9 and _PHONE_HINT_RE.search(hint):
                            masked = mask_digits_keep("0" + digits, 4, 2)
                            cell.value = masked
                            report.add(loc, MaskedItem(
                                "phone_numeric", "電話(數值)", "0" + digits, masked))
                    continue

                if not isinstance(v, str) or not v:
                    continue
                if v.startswith("="):
                    if engine.scan(v):
                        formula_hits += 1
                    continue
                new_value, items = engine.mask_text(v, context_hint=hint)
                if items:
                    report.add_all(loc, items)
                    cell.value = new_value
        if formula_hits:
            report.warn("工作表「%s」有 %d 個公式內含疑似個資，為避免破壞計算未自動改寫，請人工確認"
                        % (ws.title, formula_hits))

    if not keep_metadata:
        props = wb.properties
        if props.creator or props.lastModifiedBy:
            report.warn("已清除活頁簿屬性中的作者／最後修改者資訊")
        props.creator = ""
        props.lastModifiedBy = ""

    wb.save(output_path)
