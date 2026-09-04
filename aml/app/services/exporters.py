"""報表匯出。

年度報表版面完全依照主管機關 115 年格式重建（欄列座標對照原始 .xls），
以確保申報時可直接沿用，不需人工搬移。
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import Assessment, AssessmentStatus, RiskLevel
from ..security import decrypt_pii, mask_id_number, mask_name
from .aggregate import Dimension, PeriodSummary, risk_band

_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # 原表以黃框標示須填列之件數欄位
_TITLE_FONT = Font(bold=True, size=12)
_BOLD = Font(bold=True)
_THIN = Side(style="thin", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_PCT = "0.00%"
_NUM2 = "0.00"
_MONEY = "#,##0"

STATUS_LABELS = {
    AssessmentStatus.DRAFT: "填寫中",
    AssessmentStatus.SUBMITTED: "已送出（一般風險）",
    AssessmentStatus.PENDING_APPROVAL: "待主管同意（高風險）",
    AssessmentStatus.APPROVED: "主管已同意",
    AssessmentStatus.REJECTED: "主管不同意",
    AssessmentStatus.BLOCKED: "系統擋件（應婉拒）",
    AssessmentStatus.CLOSED: "結案",
}
LEVEL_LABELS = {RiskLevel.HIGH: "高風險", RiskLevel.GENERAL: "一般風險"}


def _put(ws, coord: str, value, *, font=None, fill=None, fmt=None, align=None, border=True):
    cell = ws[coord]
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        cell.border = _BORDER
    return cell


def _block(ws, *, title_cell: str, title: str, first_col: int, header_row: int,
           dim: Dimension, premium_rows: bool = False) -> None:
    """輸出一個風險構面區塊：標題／桶別／權重／件數／占比／平均風險分數／風險等級。

    版面依原表：標題列上方一列，接著
      header_row      : 合計件數 | 各桶別名稱
      header_row + 1  : 總件數   | 各桶別權重
      header_row + 2  : 件數     | 各桶別件數
      header_row + 3  : 占比     | 各桶別占比
      header_row + 4  : 平均風險分數 | 值 | 風險等級 | 值
    """
    def col(offset: int) -> str:
        return get_column_letter(first_col + offset)

    _put(ws, title_cell, title, font=_TITLE_FONT, fill=_HEADER_FILL, align="left")

    _put(ws, f"{col(0)}{header_row}", "合計件數", font=_BOLD, fill=_HEADER_FILL, align="center")
    _put(ws, f"{col(0)}{header_row + 1}", dim.total_count, font=_BOLD, align="center")
    _put(ws, f"{col(0)}{header_row + 2}", "件數", font=_BOLD, align="center")
    _put(ws, f"{col(0)}{header_row + 3}", "占比", font=_BOLD, align="center")

    for i, bucket in enumerate(dim.buckets, start=1):
        _put(ws, f"{col(i)}{header_row}", bucket.label,
             font=_BOLD, fill=_HEADER_FILL, align="center")
        _put(ws, f"{col(i)}{header_row + 1}", bucket.weight, align="center")
        _put(ws, f"{col(i)}{header_row + 2}", bucket.count, fill=_INPUT_FILL, align="center")
        _put(ws, f"{col(i)}{header_row + 3}", dim.share(bucket), fmt=_PCT, align="center")

    avg_row = header_row + 4
    _put(ws, f"{col(0)}{avg_row}", "平均風險分數", font=_BOLD, align="center")
    _put(ws, f"{col(1)}{avg_row}", round(dim.average_score, 2), fmt=_NUM2, align="center")
    _put(ws, f"{col(2)}{avg_row}", "風險等級", font=_BOLD, align="center")
    _put(ws, f"{col(3)}{avg_row}", risk_band(dim.average_score), align="center")

    _put(
        ws, f"{col(0)}{avg_row + 1}",
        f"說明： 上列黃框內{len(dim.buckets)}個欄位件數加總數， 需與【招攬新契約總數】相等。",
        align="left",
    )


def build_annual_workbook(summary: PeriodSummary, company_name: str) -> Workbook:
    """年度洗錢及資恐風險評估彙總表（比照主管機關 115 年格式）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "保費試算"

    widths = (22, 14, 14, 14, 14, 3, 22, 14, 14, 16, 14, 20, 16)
    for c, width in zip("ABCDEFGHIJKLM", widths, strict=True):
        ws.column_dimensions[c].width = width

    _put(ws, "A1", "招攬新契約總數", font=_TITLE_FONT, fill=_HEADER_FILL, align="center")
    _put(ws, "B1", summary.total_cases, font=_TITLE_FONT, align="center")
    _put(ws, "C1", f"統計期間　{summary.period_start:%Y-%m-%d} ～ {summary.period_end:%Y-%m-%d}",
         align="left", border=False)
    _put(ws, "A2", company_name, font=_BOLD, align="left", border=False)

    d = summary.dimensions

    # 左欄：地域風險、職業、自然人別、國籍、來源
    _block(ws, title_cell="A5", title="地域風險", first_col=1, header_row=6, dim=d["geo"])
    _block(ws, title_cell="A13", title="客戶風險 - 職業", first_col=1, header_row=14,
           dim=d["occupation"])
    _block(ws, title_cell="A21", title="客戶風險 - 自然人/非自然人", first_col=1, header_row=22,
           dim=d["entity"])
    _block(ws, title_cell="A29", title="客戶風險 - 國籍", first_col=1, header_row=30,
           dim=d["nationality"])
    _block(ws, title_cell="A37", title="客戶風險 - 來源(與業務員如何認識)", first_col=1,
           header_row=38, dim=d["source"])

    # 右欄：產品風險（件數與保費雙重加權）
    product = d["product"]
    _put(ws, "G5", "產品風險", font=_TITLE_FONT, fill=_HEADER_FILL, align="left")
    _put(ws, "G6", "合計件數", font=_BOLD, fill=_HEADER_FILL, align="center")
    _put(ws, "G7", product.total_count, font=_BOLD, align="center")
    _put(ws, "G8", "件數", font=_BOLD, align="center")
    _put(ws, "G9", "占比", font=_BOLD, align="center")
    _put(ws, "G11", "保費金額", font=_BOLD, align="center")
    _put(ws, "G12", "占比", font=_BOLD, align="center")
    for i, bucket in enumerate(product.buckets, start=1):
        col = get_column_letter(7 + i)
        _put(ws, f"{col}6", bucket.label, font=_BOLD, fill=_HEADER_FILL, align="center")
        _put(ws, f"{col}7", bucket.weight, align="center")
        _put(ws, f"{col}8", bucket.count, fill=_INPUT_FILL, align="center")
        _put(ws, f"{col}9", product.share(bucket), fmt=_PCT, align="center")
        _put(ws, f"{col}11", round(bucket.premium), fill=_INPUT_FILL, fmt=_MONEY, align="center")
        _put(ws, f"{col}12", product.premium_share(bucket), fmt=_PCT, align="center")

    count_score = product.average_score
    premium_score = product.premium_weighted_score
    _put(ws, "G10", "件數風險分數", font=_BOLD, align="center")
    _put(ws, "H10", round(count_score, 2), fmt=_NUM2, align="center")
    _put(ws, "I10", "風險等級", font=_BOLD, align="center")
    _put(ws, "J10", risk_band(count_score), align="center")
    _put(ws, "G13", "保費風險分數", font=_BOLD, align="center")
    _put(ws, "H13", round(premium_score, 2), fmt=_NUM2, align="center")
    _put(ws, "I13", "風險等級", font=_BOLD, align="center")
    _put(ws, "J13", risk_band(premium_score), align="center")
    combined = (count_score + premium_score) / 2
    _put(ws, "G14", "產品風險平均分數", font=_BOLD, align="center")
    _put(ws, "H14", round(combined, 2), fmt=_NUM2, align="center")
    _put(ws, "I14", "風險等級", font=_BOLD, align="center")
    _put(ws, "J14", risk_band(combined), align="center")
    _put(
        ws, "G15",
        "說明： 上列表單請分別填入【件數】及【金額】，"
        "五個欄位件數加總數，需與【招攬新契約總數】相等。",
        align="left",
    )

    _put(ws, "M10", "新契約保費", font=_BOLD, fill=_HEADER_FILL, align="center")
    _put(ws, "M11", round(summary.total_premium), fmt=_MONEY, align="center")

    # 右欄：交易風險
    _block(ws, title_cell="G17", title="交易風險", first_col=7, header_row=18, dim=d["channel"])

    # 右欄：制裁名單、STR 與境外電匯件數（原表為獨立計數區，無權重）
    _put(ws, "G25", "制裁名單、STR與境外電匯件數(含OIU保單)", font=_TITLE_FONT,
         fill=_HEADER_FILL, align="left")
    _put(ws, "G26", "合計件數", font=_BOLD, fill=_HEADER_FILL, align="center")
    counters = [
        ("制裁名單", summary.sanction_hits),
        ("STR", summary.str_reported_cases),
        ("境外電匯件數(含OIU保單)", summary.offshore_remittance_cases),
    ]
    _put(ws, "G27", sum(v for _, v in counters), font=_BOLD, align="center")
    _put(ws, "G28", "件數", font=_BOLD, align="center")
    _put(ws, "G29", "占比", font=_BOLD, align="center")
    for i, (label, value) in enumerate(counters, start=1):
        col = get_column_letter(7 + i)
        _put(ws, f"{col}27", label, font=_BOLD, fill=_HEADER_FILL, align="center")
        _put(ws, f"{col}28", value, fill=_INPUT_FILL, align="center")
        share = value / summary.total_cases if summary.total_cases else 0.0
        _put(ws, f"{col}29", share, fmt=_PCT, align="center")
    _put(ws, "G30",
         "說明：本區為件數統計，占比之分母為招攬新契約總數；原表「平均風險分數」欄位未定義權重，"
         "如主管機關另有計算方式，請於申報前依函令調整。",
         align="left")

    return wb


def _sheet_header(ws, headers: list[str], widths: list[int]) -> None:
    for i, (h, w) in enumerate(zip(headers, widths, strict=True), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        _put(ws, f"{get_column_letter(i)}1", h, font=_BOLD, fill=_HEADER_FILL, align="center")
    ws.freeze_panes = "A2"


def build_board_workbook(
    summary: PeriodSummary, cases: list[Assessment], company_name: str,
    org_names: dict[int, str],
) -> Workbook:
    """董事會定期報告：總覽、高風險與擋件明細、單位別分布。"""
    wb = Workbook()

    ws = wb.active
    ws.title = "總覽"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 52
    _put(ws, "A1", f"{company_name}　防制洗錢及打擊資恐　客戶風險評估報告",
         font=Font(bold=True, size=14), align="left")
    _put(ws, "A2", f"報告期間：{summary.period_start:%Y-%m-%d} ～ {summary.period_end:%Y-%m-%d}",
         align="left", border=False)

    rows = [
        ("招攬新契約評估總件數", summary.total_cases, "已送出之評估案件（不含填寫中草稿）"),
        ("新契約保費合計", round(summary.total_premium), "單位：新臺幣元"),
        ("一般風險件數", summary.general_risk_cases, "總分 29 分以下（含）且未觸發強制規則"),
        ("高風險件數", summary.high_risk_cases, "總分 30 分以上（含），或經強制規則提升"),
        ("高風險占比", f"{summary.high_risk_ratio:.2%}", "高風險件數 ÷ 評估總件數"),
        ("待主管同意件數", summary.pending_approval_cases,
         "範本第五點：高風險應於建立業務關係前取得高階管理人員同意"),
        ("主管已同意件數", summary.approved_high_risk, "已完成強化確認客戶身分措施"),
        ("主管不同意件數", summary.rejected_cases, ""),
        ("系統擋件（應婉拒）件數", summary.blocked_cases, "範本第四點所定應婉拒建立業務關係之情形"),
        ("經強制規則提升為高風險件數", summary.override_cases,
         "制裁名單地域、PEP、高風險國家、疑似洗錢態樣等"),
        ("命中制裁名單件數", summary.sanction_hits, ""),
        ("已申報疑似洗錢交易（STR）件數", summary.str_reported_cases,
         "範本第九點：不論金額多寡均應申報"),
        ("境外電匯／OIU 保單件數", summary.offshore_remittance_cases, ""),
        ("尚未送出之草稿件數", summary.draft_cases, "期末仍在填寫中，應追蹤是否逾期未完成"),
    ]
    for i, (label, value, note) in enumerate(rows, start=4):
        _put(ws, f"A{i}", label, font=_BOLD, align="left")
        _put(ws, f"B{i}", value, align="center")
        _put(ws, f"C{i}", note, align="left")

    start = len(rows) + 6
    _put(ws, f"A{start}", "各風險構面平均分數", font=_TITLE_FONT, fill=_HEADER_FILL, align="left")
    _put(ws, f"B{start}", "平均風險分數", font=_BOLD, fill=_HEADER_FILL, align="center")
    _put(ws, f"C{start}", "風險等級", font=_BOLD, fill=_HEADER_FILL, align="center")
    for offset, dim in enumerate(summary.dimensions.values(), start=1):
        r = start + offset
        _put(ws, f"A{r}", dim.label, align="left")
        _put(ws, f"B{r}", round(dim.average_score, 2), fmt=_NUM2, align="center")
        _put(ws, f"C{r}", risk_band(dim.average_score), align="center")

    # 高風險與擋件明細
    detail = wb.create_sheet("高風險及擋件明細")
    _sheet_header(
        detail,
        ["案號", "送件日", "業務員", "所屬單位", "要保人", "身分證字號", "保險公司",
         "保費", "總分", "風險等級", "狀態", "強制規則／擋件事由"],
        [18, 12, 12, 16, 12, 14, 16, 12, 8, 10, 20, 60],
    )
    flagged = [
        c for c in cases
        if c.risk_level == RiskLevel.HIGH or c.status == AssessmentStatus.BLOCKED
    ]
    import json as _json

    for r, case in enumerate(flagged, start=2):
        reasons: list[str] = []
        for raw in (case.override_reasons, case.blocked_reasons):
            if raw:
                try:
                    reasons.extend(_json.loads(raw))
                except _json.JSONDecodeError:
                    reasons.append(raw)
        values = [
            case.case_no,
            case.submitted_at.strftime("%Y-%m-%d") if case.submitted_at else "",
            case.agent.display_name if case.agent else "",
            org_names.get(case.org_unit_id, ""),
            mask_name(decrypt_pii(case.holder_name_enc)),
            mask_id_number(decrypt_pii(case.holder_id_enc)),
            case.insurer_name or "",
            case.annual_premium or 0,
            case.total_score,
            LEVEL_LABELS.get(case.risk_level, ""),
            STATUS_LABELS.get(case.status, ""),
            "；".join(reasons),
        ]
        for i, v in enumerate(values, start=1):
            _put(detail, f"{get_column_letter(i)}{r}", v,
                 fmt=_MONEY if i == 8 else None, align="left")

    # 單位別分布
    org_sheet = wb.create_sheet("單位別分布")
    _sheet_header(org_sheet, ["所屬單位", "評估件數", "高風險件數", "高風險占比", "擋件件數"],
                  [24, 12, 12, 12, 12])
    for r, (org_id, stats) in enumerate(sorted(summary.by_org.items()), start=2):
        name = org_names.get(int(org_id), "未指定") if org_id.isdigit() else "未指定"
        ratio = stats["high"] / stats["total"] if stats["total"] else 0.0
        row = [name, stats["total"], stats["high"], ratio, stats["blocked"]]
        for i, v in enumerate(row, start=1):
            _put(org_sheet, f"{get_column_letter(i)}{r}", v,
                 fmt=_PCT if i == 4 else None, align="center" if i > 1 else "left")

    return wb


def build_case_register(cases: list[Assessment], org_names: dict[int, str],
                        unmask: bool = False) -> Workbook:
    """案件清冊。供內部稽核與金融檢查調閱，預設遮罩個資。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "案件清冊"
    _sheet_header(
        ws,
        ["案號", "建立日", "送件日", "業務員", "所屬單位", "要保人", "身分證字號",
         "保險公司", "保單號碼", "保費", "總分", "風險等級", "狀態", "已申報STR"],
        [18, 12, 12, 12, 16, 14, 16, 16, 16, 12, 8, 10, 20, 10],
    )
    for r, case in enumerate(cases, start=2):
        name = decrypt_pii(case.holder_name_enc)
        idno = decrypt_pii(case.holder_id_enc)
        values = [
            case.case_no,
            case.created_at.strftime("%Y-%m-%d") if case.created_at else "",
            case.submitted_at.strftime("%Y-%m-%d") if case.submitted_at else "",
            case.agent.display_name if case.agent else "",
            org_names.get(case.org_unit_id, ""),
            (name or "") if unmask else mask_name(name),
            (idno or "") if unmask else mask_id_number(idno),
            case.insurer_name or "",
            case.policy_no or "",
            case.annual_premium or 0,
            case.total_score,
            LEVEL_LABELS.get(case.risk_level, ""),
            STATUS_LABELS.get(case.status, ""),
            "是" if case.str_reported else "否",
        ]
        for i, v in enumerate(values, start=1):
            _put(ws, f"{get_column_letter(i)}{r}", v, fmt=_MONEY if i == 10 else None, align="left")
    return wb
