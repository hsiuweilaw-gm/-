"""制裁名單匯入。

支援「制裁名單篩查系統」產出的 .xlsx，以及同樣欄位的 .csv。
欄位：來源清單、編號、類型、姓名/名稱(原文)、中文、別名、國家/地區、
      依據/計畫、列管日期、狀態、備註。

匯入以「批次」為單位：同一來源清單的舊資料在新批次匯入後停用，
避免已從名單移除的對象仍留在系統中比對。停用而非刪除，以保留曾經比對之依據。
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import UTC, datetime

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ..models import WatchListEntry, WatchListName
from . import screening

COLUMNS = {
    "source": ("來源清單", "source"),
    "external_id": ("編號", "id"),
    "entity_type": ("類型", "type"),
    "name": ("姓名/名稱(原文)", "姓名/名稱", "name"),
    "name_zh": ("中文", "name_zh"),
    "aliases": ("別名", "alias", "aliases"),
    "countries": ("國家/地區", "country"),
    "program": ("依據/計畫", "program"),
    "listed_on": ("列管日期", "listed_on"),
    "status": ("狀態", "status"),
    "note": ("備註", "note"),
}
ALIAS_SPLIT = re.compile(r"[;；]")

# 資恐防制法第 4 條指定制裁者列入資恐名單，其餘來源列為制裁名單。
TERRORIST_MARKERS = ("資恐防制法", "SDGT", "AL-QAIDA", "TALIBAN", "ISIL")


@dataclass
class ImportResult:
    batch: str
    entries: int = 0
    names: int = 0
    deactivated: int = 0
    skipped: int = 0
    by_source: dict[str, int] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


def _header_map(header: list[str]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for i, raw in enumerate(header):
        cleaned = str(raw or "").strip().lstrip("﻿").lower()
        for key, aliases in COLUMNS.items():
            if cleaned in {a.lower() for a in aliases}:
                mapping[i] = key
                break
    return mapping


def _classify(source: str, program: str) -> str:
    haystack = f"{source} {program}".upper()
    return "terrorist" if any(m in haystack for m in TERRORIST_MARKERS) else "sanction"


def rows_from_xlsx(data: bytes, sheet: str | None = None):
    """讀出 xlsx 的資料列。預設取「全部清單」，找不到則取第一張含資料的工作表。"""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    names = wb.sheetnames
    target = sheet or ("全部清單" if "全部清單" in names else names[0])
    ws = wb[target]
    for row in ws.iter_rows(values_only=True):
        yield ["" if v is None else str(v) for v in row]


def rows_from_csv(text: str):
    normalized = text.lstrip("﻿").replace("\r\n", "\n").replace("\r", "\n")
    yield from csv.reader(io.StringIO(normalized))


def import_rows(db: Session, rows, *, batch: str | None = None,
                replace_sources: bool = True) -> ImportResult:
    """匯入名單。replace_sources 為真時，本批出現過的來源，其舊資料一律停用。"""
    batch = batch or datetime.now(UTC).strftime("%Y%m%d%H%M%S")
    result = ImportResult(batch=batch)

    iterator = iter(rows)
    try:
        header = next(iterator)
    except StopIteration:
        result.errors.append("檔案是空的")
        return result

    mapping = _header_map(list(header))
    if "name" not in mapping.values():
        result.errors.append("找不到「姓名/名稱(原文)」欄位，請確認檔案格式")
        return result

    seen_sources: set[str] = set()
    for row in iterator:
        values = {key: (row[i].strip() if i < len(row) else "") for i, key in mapping.items()}
        name = values.get("name", "")
        name_zh = values.get("name_zh", "")
        if not name and not name_zh:
            result.skipped += 1
            continue

        source = values.get("source") or "未標示"
        seen_sources.add(source)
        aliases = [a.strip() for a in ALIAS_SPLIT.split(values.get("aliases", "")) if a.strip()]
        entry = screening.add_entry(
            db,
            _classify(source, values.get("program", "")),
            name or name_zh,
            source=source,
            external_id=values.get("external_id") or None,
            name_zh=name_zh or None,
            aliases=aliases,
            entity_type=values.get("entity_type") or None,
            countries=values.get("countries") or None,
            program=values.get("program") or None,
            listed_on=values.get("listed_on") or None,
            status=values.get("status") or None,
            note=values.get("note") or None,
            batch=batch,
        )
        if entry is None:
            result.skipped += 1
            continue
        result.entries += 1
        result.names += len(entry.names)
        result.by_source[source] = result.by_source.get(source, 0) + 1
        if result.entries % 2000 == 0:
            db.flush()

    db.flush()
    if replace_sources and seen_sources:
        stale = (
            db.query(WatchListEntry)
            .filter(WatchListEntry.source.in_(seen_sources),
                    WatchListEntry.batch != batch,
                    WatchListEntry.active.is_(True))
            .all()
        )
        for old in stale:
            old.active = False
        result.deactivated = len(stale)
    return result


def summary(db: Session) -> dict:
    """目前生效名單的統計，供名單管理頁呈現。"""
    by_source = dict(
        db.execute(
            select(WatchListEntry.source, func.count(WatchListEntry.id))
            .where(WatchListEntry.active.is_(True))
            .group_by(WatchListEntry.source)
        ).all()
    )
    by_type = dict(
        db.execute(
            select(WatchListEntry.list_type, func.count(WatchListEntry.id))
            .where(WatchListEntry.active.is_(True))
            .group_by(WatchListEntry.list_type)
        ).all()
    )
    names = db.execute(
        select(func.count(WatchListName.id))
        .join(WatchListEntry)
        .where(WatchListEntry.active.is_(True))
    ).scalar_one()
    latest = db.execute(
        select(func.max(WatchListEntry.created_at)).where(WatchListEntry.active.is_(True))
    ).scalar_one()
    return {"by_source": by_source, "by_type": by_type, "names": names,
            "entries": sum(by_source.values()), "updated_at": latest}
